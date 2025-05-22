import pandas as pd
import pyautogui
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

dados = {
    'Tarefa': ['abrir windows', 'pesquisar chrome', 'abrir o chrome', 'espera', 'pesquisar youtube', 'enter', 'espera', 'pesquisar musica', 'enter', 'espera', 'clicar no video', 'espera'],
    'Tipo': ['click', 'texto', 'click', 'espera', 'texto', 'tecla', 'espera', 'texto', 'tecla', 'espera', 'click', 'espera'],
    'Dado': ['win', 'chrome', 'chrome_icone', 2, 'https://www.youtube.com/', 'enter', 2, 'over the moon', 'enter', 2, 'clicar no video', 2],
}

df_tarefas = pd.DataFrame(dados)

arquivo_tarefas = 'ListaTarefas.xlsx'
df_tarefas.to_excel(arquivo_tarefas, index=False, sheet_name='Tarefas')

wb = load_workbook(arquivo_tarefas)
ws = wb['Tarefas']


for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save(arquivo_tarefas)
print(f"Planilha '{arquivo_tarefas}' criada e formatada com sucesso.")

pyautogui.PAUSE = 2

relatorio_dados = {
    'Tarefa': df_tarefas['Tarefa'],
    'Tempo Início': [None] * len(df_tarefas),
    'Tempo Fim': [None] * len(df_tarefas),
    'Tempo de Execução (s)': [None] * len(df_tarefas),
    'Status': ['Pendente'] * len(df_tarefas),
}

df_relatorio = pd.DataFrame(relatorio_dados)

arquivo_relatorio = 'RelatorioExecucao.xlsx'
df_relatorio.to_excel(arquivo_relatorio, index=False, sheet_name='Relatorio')

def registrar_tempo(i, acao):
    inicio = time.time()
    try:
        acao()
        fim = time.time()

        tempo_execucao = round(fim - inicio, 2)
        
        df_relatorio.at[i, 'Tempo Início'] = round(inicio, 2)
        df_relatorio.at[i, 'Tempo Fim'] = round(fim, 2)
        df_relatorio.at[i, 'Tempo de Execução (s)'] = tempo_execucao
        df_relatorio.at[i, 'Status'] = 'Concluída'

    except Exception as e:
        fim = time.time()
        tempo_execucao = round(fim - inicio, 2)
        df_relatorio.at[i, 'Tempo Início'] = round(inicio, 2)
        df_relatorio.at[i, 'Tempo Fim'] = round(fim, 2)
        df_relatorio.at[i, 'Tempo de Execução (s)'] = tempo_execucao
        df_relatorio.at[i, 'Status'] = f'Erro: {str(e)}'
    df_relatorio.to_excel(arquivo_relatorio, index=False, sheet_name='Relatorio')



def abrirChrome():
    pyautogui.press('win')
    pyautogui.write('chrome', interval=0.1)
    pyautogui.press('enter')
    time.sleep(2)

def abrirYoutube():
    pyautogui.write('https://www.youtube.com')
    pyautogui.press('enter')
    time.sleep(2)

def pesquisarMusica(musica):
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    time.sleep(2)
    pyautogui.write(musica)
    pyautogui.press('enter')
    time.sleep(2)
    pyautogui.click(x=713, y=778)
    time.sleep(2)

def automacao_principal():
    registrar_tempo(0, abrirChrome)
    registrar_tempo(1, abrirYoutube)
    musica = 'Over the Moon'
    registrar_tempo(2, lambda: pesquisarMusica(musica))
    registrar_tempo(3, lambda: time.sleep(2))
    registrar_tempo(4, lambda: time.sleep(2))
    registrar_tempo(5, lambda: pyautogui.press('enter'))
    registrar_tempo(6, lambda: time.sleep(2))
    registrar_tempo(7, lambda: pesquisarMusica('over the Moon'))
    registrar_tempo(8, lambda: pyautogui.press('enter'))
    registrar_tempo(9, lambda: time.sleep(2))
    registrar_tempo(10, lambda: pyautogui.click(x=713, y=778))
    registrar_tempo(11, lambda: time.sleep(2))


automacao_principal()


wb = load_workbook(arquivo_relatorio)
ws = wb['Relatorio']


verde_claro = PatternFill(start_color='90EE90', fill_type='solid')
vermelho = PatternFill(start_color='ff3333', fill_type='solid')


for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        if cell.value == 'Concluída':
            cell.fill = verde_claro
        else:
            cell.fill = vermelho 


wb.save(arquivo_relatorio)